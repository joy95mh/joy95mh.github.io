from flask import Flask, render_template, request, send_file, make_response, jsonify
import psycopg2
import openpyxl
from openpyxl.utils import get_column_letter
import io
import re, scrapy

from flask_socketio import SocketIO, emit

app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret!'
socketio = SocketIO(app, async_mode='threading')


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Get the form data
        domain_netloc = request.form['domain_netloc']
        start_link = request.form['start_link']
        xcategory_link = request.form['xcategory_link']
        xproduct_link = request.form['xproduct_link']
        xnext_link = request.form['xnext_link']
        xname = request.form['xname']
        xprice = request.form['xprice']
        xprice_online = request.form['xprice_online']
        ximage = request.form['ximage']
        xdescription = request.form['xdescription']
        xdescription_content = request.form['xdescription_content']

        # Connect to the database
        conn = psycopg2.connect(
            host="192.168.101.236",
            port='5432',
            database="lazada",
            user="postgres",
            password="123456",
            
        )

        # Insert the data into the database
        cur = conn.cursor()
        conn.autocommit = True
        query = '''
            insert into skus_xpath(domain_netloc, start_link, xcategory_link, 
            xproduct_link, xnext_link, xname, xprice, xprice_online, ximage, xdescription, 
            xdescription_content)
            values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s') on conflict (domain_netloc) do
            update set
            start_link = EXCLUDED.start_link
            ,xcategory_link = EXCLUDED.xcategory_link
            ,xproduct_link = EXCLUDED.xproduct_link
            ,xnext_link = EXCLUDED.xnext_link
            ,xname = EXCLUDED.xname
            ,xprice = EXCLUDED.xprice
            ,xprice_online = EXCLUDED.xprice_online
            ,ximage = EXCLUDED.ximage
            ,xdescription = EXCLUDED.xdescription
            ,xdescription_content = EXCLUDED.xdescription_content
        '''
        
        xpath = {
            'domain_netloc':domain_netloc,
            'start_link':start_link,
            'xcategory_link':xcategory_link,
            'xproduct_link':xproduct_link,
            'xnext_link':xnext_link,
            'xname':xname,
            'xprice':xprice,
            'xprice_online':xprice_online,
            'ximage':ximage,
            'xdescription':xdescription,
            'xdescription_content':xdescription_content,
            }

        #query = query==query_3 and query%services_category or query
        try:
            cur.execute(query%(
                xpath.get('domain_netloc')
                ,xpath.get('start_link')
                ,xpath.get('xcategory_link')
                ,xpath.get('xproduct_link')
                ,xpath.get('xnext_link')
                ,xpath.get('xname')
                ,xpath.get('xprice')
                ,xpath.get('xprice_online')
                ,xpath.get('ximage')
                ,xpath.get('xdescription')
                ,xpath.get('xdescription_content')
            ))
        except Exception as e:
            #print ('ERROR:',re.search('(?<=CONTEXT:  COPY tmp_table,).*',str(e),re.I).group())
            print(e)
            #sys.exit(0)
            
        query_delete = '''
            delete from skus where domain_netloc='%s'
            '''%xpath.get('domain_netloc')

        try:
            cur.execute(query_delete)#%file
        except Exception as e:
            #print ('ERROR:',re.search('(?<=CONTEXT:  COPY tmp_table,).*',str(e),re.I).group())
            print(e)
            #sys.exit(0)

        cur.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'skus_xpath'")
        column_names = [row[0] for row in cur.fetchall()]
        cur.execute("SELECT * FROM skus_xpath where domain_netloc = '%s'"%(xpath.get('domain_netloc')) )
        data = [row for row in cur.fetchall()]

        conn.close()

        conn.close()
        return render_template('skus-run.html',column_names=column_names,data=data)

    return render_template('skus-run.html')

@app.route('/skus-search')
def skus_search():
    limit = request.args.get('limit', default = 100, type = int)
    page = request.args.get('page', default = 1, type = int)
    offset = (page - 1) * limit
    filter1 = request.args.get('filter1', '').strip()
    conn = psycopg2.connect(
        host="192.168.101.236",
        port='5432',
        database="lazada",
        user="postgres",
        password="123456",
        
    )
    c = conn.cursor()
    c.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'skus'")
    column_names = [row[0] for row in c.fetchall()]
    if limit > 0 and filter1 != '':
       c.execute("SELECT count(*) FROM skus where domain_netloc = '%s'"%(filter1))
       count_all = [row[0] for row in c.fetchall()][0]
       max_page = int(count_all / float(limit)) + 1
       pages = [p for p in range(max(1, page-2), min(max_page+1, page+3))]
       c.execute("SELECT * FROM skus where domain_netloc = '%s' or product_url='%s' ORDER BY product_id asc LIMIT %s OFFSET %s"%(filter1,filter1,limit,offset))
    else:
        pages = []
        num_pages = 1
        max_page = 1
        count_all = 0
        c.execute('SELECT * FROM skus order by product_id desc limit %s'%(limit or 0))
        
    
    
    data = [row for row in c.fetchall()]
    has_prev = (page > 1)
    has_next = (len(data) == limit)
    conn.close()
    return render_template("skus-search.html", has_prev=has_prev,has_next=has_next,data=data, column_names=column_names, count=len(data), count_all=count_all, page=page, pages=pages, max_page=max_page, filter1=filter1, limit=limit)
@app.route('/skus-xpath')
def skusxpath():
    filter = request.args.get('filter', '').strip()
    filter1 = request.args.get('filter1', '').strip()
    conn = psycopg2.connect(
        host="192.168.101.236",
        port='5432',
        database="lazada",
        user="postgres",
        password="123456",
        
    )
    c = conn.cursor()
    c.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'skus_xpath'")
    column_names = [row[0] for row in c.fetchall()]
    if filter != '' and filter1 != '':
       c.execute("SELECT * FROM skus_xpath where domain_netloc = '%s'"%(filter1) )
    elif filter == '' and filter1 == '':
        c.execute('SELECT * FROM skus_xpath order by id desc limit 30')
    else:
        c.execute("SELECT * FROM skus_xpath where domain_netloc = '%s'"%(filter == '' and filter1 or filter) )
    
    data = [row for row in c.fetchall()]
    conn.close()
    return render_template('skus-xpath.html', column_names=column_names,data=data, filter=filter, filter1=filter1, count=len(data))#, filter_options=filter_options

@app.route('/shopee-checkrun')
def shopeecheckrun():
    daybefore = request.args.get('daybefore', '1').strip()
    filter1 = request.args.get('filter1', '').strip()
    conn = psycopg2.connect(
        host="192.168.101.236",
        port='5432',
        database="lazada",
        user="postgres",
        password="123456",
        
    )
    c = conn.cursor()
    c.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'shopee_checkrun'")
    column_names = [row[0] for row in c.fetchall()]
    if daybefore != '' and filter1 != '':
       c.execute("SELECT * FROM shopee_checkrun where domain_netloc = '%s'"%(filter1) )
    else:
        c.execute("select * from shopee_checkrun where created_date::date >= current_Date -'%s day'::interval order by created_date::date desc"%daybefore)
    
    data = [row for row in c.fetchall()]
    conn.close()
    return render_template('shopee-checkrun.html', column_names=column_names,data=data, daybefore=daybefore, filter1=filter1, count=len(data))

def validator_image(images,folder):
        
    return images[0]

@app.route('/export_to_xlsx')
def export_to_xlsx():
    filter1 = request.args.get('filter1')
    limit = request.args.get('limit', default=3200, type=int)
    filter2 = request.args.get('filter1','//div[@id="variant-swatch-1"]//div/@data-value')
    filter3 = request.args.get('filter2','//div[@id="variant-swatch-0"]//div/@data-value')

    # Execute a SQL query and get the results
    conn = psycopg2.connect(
        host="192.168.101.236",
        port='5432',
        database="lazada",
        user="postgres",
        password="123456",
    )
    cur = conn.cursor()
    cur.execute(f'''select product_id, 
                    Case when regexp_replace(price,'\D','','g') = '' then '0' 
                    when online_price ~ '-' then regexp_replace(regexp_replace(online_price,'[-].*',''),'\D','','g')
                    else 
                    regexp_replace(regexp_replace(price,'.*[-]',''),'\D','','g') end as price,
                    Case when regexp_replace(online_price,'\D','','g') = '' then '0' 
                    when online_price ~ '-' then regexp_replace(regexp_replace(online_price,'.*[-]',''),'\D','','g')
                    else 
                    regexp_replace(regexp_replace(online_price,'(?<=\d.*)[-].*',''),'\D','','g') end as online_price,
                    product_name "name",
                    short_desc, description, description_content,short_desc brand, '' size,'' color,image,
                    product_url url, domain_netloc "domain", attr "attribute",
                    created_date::timestamp(0) created_at
                    from skus where domain_netloc ~'{filter1}'
                    order by 1 LIMIT {limit}''')
    results = cur.fetchall()
    col_names = [desc[0] for desc in cur.description]

    # Create a workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    # Write the column names to the first row
    worksheet.append(col_names)
    worksheet.column_dimensions[get_column_letter(1)].width = 20
    worksheet.column_dimensions[get_column_letter(2)].width = 10
    # Write the results to the worksheet
        # Truncate cell contents if they exceed 50 characters
    folder=""
    rowid =2
    desc_xpath = '//div[text()="Số lượng"]/parent::div/parent::div/parent::div'
    for r, row in enumerate(results):
        #try:
        #images_raw = row[10]
        images_raw = row[10].encode().decode('unicode_escape')
        selector = scrapy.Selector(text=row[6])
        #brand = selector.xpath('//div[contains(.//text(),"Thương hiệu")]/span/text()|//div[contains(.//text(),"Thương Hiệu")]/div/text()|//div[contains(.//text(),"Thương hiệu")]/span/text()|//p[contains(.//text(),"Thương hiệu")]/text()[1]').extract_first() or ''
        brand = selector.xpath('//label[text()="Thương hiệu"]/following-sibling::span/text()').extract_first()
        #print(brand)
        if row[5] is not None:
            sel1 = scrapy.Selector(text=row[6] or '')


            #--------Shopee--------------------
            if 'shopee.vn' in row[11]:
                cate = ' > '.join(sel1.xpath('//label[text()="Danh Mục"]/following-sibling::div[1]//text()').extract()[1:])
                desc_content = sel1.xpath('//div[text()="MÔ TẢ SẢN PHẨM" or text()="label_product_description"]/parent::div').extract_first() or ''
                desc_content = re.sub('|\d+\.\d+\s?₫|₫\s?\d+\.\d+|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',desc_content)
            else:
                cate = ''
                desc_content = row[6]
            sel = scrapy.Selector(text=row[5])
            #sizes = sel.xpath('//div[contains(@class,"flex items-center")]//button[contains(@class,"product-variation")]/text()').extract()
            sizes = len(sel.xpath(desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")]/label[re:test(text(),"Size|Kích|kích|KÍCH|S,M,L,XL|size|SIze|SiZe|SizE|SIZE|DÒNG")]')) < 2 and\
            sel.xpath(desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")]/label[re:test(text(),"Size|Kích|kích|KÍCH|S,M,L,XL|size|SIze|SiZe|SizE|SIZE|DÒNG")]/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract() or \
            sel.xpath('('+desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")]/label[re:test(text(),"Size|Kích|kích|KÍCH|S,M,L,XL|size|SIze|SiZe|SizE|SIZE|DÒNG")])[2]/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract()
            colors = len(sel.xpath(desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")]/label[re:test(text(),"Màu|MÀU|Colo|màu|mÀu")]')) < 2 and \
            sel.xpath(desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")]/label[re:test(text(),"Màu|MÀU|Colo|màu|mÀu")]/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract() or \
            sel.xpath('('+desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")]/label[re:test(text(),"Màu|MÀU|Colo|màu|mÀu")])[1]/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract()             
            if sizes == colors:
                sizes = []
                colors = []
            if len(sel.xpath(desc_xpath+'/div[@class="flex flex-column"]/div[@class="flex items-center"]')) < 2:
                if len(sizes) > 0:
                    colors = []
                elif len(colors) > 0:     
                    sizes = []
                else:
                    sizes = sel.xpath(desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")]/label/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract()
                    colors = []
            else:#length selection == 2   
                if len(sizes) == 0:
                    if len(colors) == 0:
                        #sizes = sel.xpath('('+desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center _90fTvx")])[re:test(.//button/text(),"\d")]/label/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract()
                        #if len(sizes) == 0:
                        sizes = sel.xpath('('+desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")])[2]/label/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract()
                        colors = sel.xpath('('+desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")])[1]/label/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract()
                        #else:    
                        #    colors = sel.xpath('('+desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center _90fTvx")])[not(re:test(.//button/text(),"\d"))]/label/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract()
                    else:
                        sizes = sel.xpath(desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")]/label[not(re:test(text(),"Màu|MÀU|Colo|màu|mÀu"))]/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract()
                elif len(colors) == 0:  
                    colors = sel.xpath(desc_xpath+'/div[@class="flex flex-column"]/div[not(@class="flex items-center ")]/label[not(re:test(text(),"Size|Kích|kích|KÍCH|S,M,L,XL|size|SIze|SiZe|SizE|SIZE|DÒNG"))]/following-sibling::div/button[contains(@class,"product-variation")]/text()').extract()



            #--------domain thường nếu phát hiện sizes + colors---------------------------------------

            if len(sizes) == 0 and len(colors) == 0:                                                #-                                                              
                #colors = [x.strip() for x in sel.xpath('//div[contains(@class,"n-sd swatch-element color")]/input/@value').extract() if x.strip()]  #-
                #sizes = [x.strip() for x in sel.xpath('//div[contains(@class,"n-sd swatch-element ")][not(contains(@class,"n-sd swatch-element color"))]/input/@value').extract() if x.strip()]  #-
                colors = sel.xpath('//div[@id="variant-swatch-0"]//div/@data-value').extract()
                sizes = sel.xpath('//div[@id="variant-swatch-0"]//div/@data-value' ).extract()
            #-----------------------------------------------------------------------------------------




            #print('Sizes: ',sizes,'Colors: ',colors,' ',row[11],len(images_raw))
            #row[11] = re.sub('(?<=shopee.vn/)\W+','',row[11])
            if len(sizes)>0:
                for i,size in enumerate(sizes):
                    if len(colors) > 0:
                        for ic,color in enumerate(colors):
                            for c, col in enumerate(row):
                                if col == 11: col = 'https://shopee.vn/%s%s'%(re.sub('\W+','-',row[3]),re.sub('.*(?=-i\.)','',col))
                                if c == 13: col = cate
                                if c == 1: col=max(eval(row[1] or '0'),eval(row[2] or '0'))
                                if c == 2: col=min(eval(row[1] or '0'),eval(row[2] or '0'))    
                                #if c == 10: 
                                #    col=len(eval(col))>0 and validator_image(eval(col),folder) or None
                                #    for ii,x in enumerate(eval(row[10])):
                                #        x=re.sub('_tn$','',x)
                                #        _ = worksheet.cell(column=c+1+5+ii, row=r+rowid+i, value=x)
                                #if c == 10 and col is not None:  url_check(col) and True or print(col)
                                if c == 10:
                                    images1 = [re.sub('_tn$','',x) for x in eval(re.sub('background-image: url\(\"|\);[^\"]+\"','',re.sub(r'\\','',images_raw))) if x.strip()]
                                    col = len(eval(row[10]))>0 and validator_image(images1,folder) or None
                                    for ii,x in enumerate(len(images1) > 5 and images1[2:] or images1[1:]):
                                        x=re.sub('_tn$','',x)
                                        _ = worksheet.cell(column=c+1+5+ii, row=r+rowid+i+ic, value=x)
                                #if c == 10 and col is not None:  url_check(col) and True or print(col)
                                if c == 8: col = size
                                if c == 9: col = color
                                if c ==4: col = ''
                                if c ==7: col = brand
                                if c== 6: col = re.sub('[\d\.]+\.000\s?|[\d\.]+\s?VNĐ|[\d\.]+\s?VND|\d+\.\d{3}\s?đ|\d+\.\d{3}\s?₫|₫\s?\d+\.\d{3}|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',desc_content)
                                if c == 5: col = re.sub('|[\d\.]+\.000\s?|[\d\.]+\s?VNĐ|[\d\.]+\s?VND|\d+\.\d+\s?đ|\d+\.\d+\s?₫|₫\s?\d+\.\d+|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',col)
                                #if c == 3: col = re.sub('\(\)','',re.sub('\d+\.\d+\s?₫|₫\s?\d+\.\d+|\d+k\/?|\d+K\/?|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',col)) 
                                _ = worksheet.cell(column=c+1, row=r+rowid+i+ic, value=col)
                        rowid += len(colors) - 1
                    else:  
                        color = ''
                        for c, col in enumerate(row):
                            if col == 11: col = 'https://shopee.vn/%s%s'%(re.sub('\W+','-',row[3]),re.sub('.*(?=-i\.)','',col))
                            if c == 13: col = cate
                            #if not re.search('size|\d',size,re.I):#Màu\s[\w\s]+|
                            #    if re.search('^Trắng$|^Hồng$|^Đỏ$|^Xanh$|^xanh ngọc$|^xanh rêu$|^xanh da trời$|^xanh dương$|^xanh cốm%|^Vàng$|^Nâu$|^Đen$|^Be$|^Chấm Bi$|^Bạc$|^Xám$|^Họa Tiết$|^Tím$',size,re.I):
                            #        color = size
                            #        size = ''
                            if c == 10:
                                images1 = [re.sub('_tn$','',x) for x in eval(re.sub('background-image: url\(\"|\);[^\"]+\"','',re.sub(r'\\','',images_raw))) if x.strip()]
                                col = len(eval(row[10]))>0 and validator_image(images1,folder) or None
                                for ii,x in enumerate(len(images1) > 5 and images1[2:] or images1[1:]):
                                    x=re.sub('_tn$','',x)
                                    _ = worksheet.cell(column=c+1+5+ii, row=r+rowid+i, value=x)
                            #if c == 10 and col is not None:  url_check(col) and True or print(col)
                            if c == 1: col=max(eval(row[1] or '0'),eval(row[2] or '0'))
                            if c == 2: col=min(eval(row[1] or '0'),eval(row[2] or '0'))    
                            if c == 9: col=color
                            if c ==4: col = ''
                            if c ==7: col = brand
                            if c == 8: col = size
                            if c == 5: col = re.sub('|[\d\.]+\.000\s?|[\d\.]+\s?VNĐ|[\d\.]+\s?VND|\d+\.\d+\s?đ|\d+\.\d+\s?₫|₫\s?\d+\.\d+|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',col)    
                            if c== 6: col = re.sub('[\d\.]+\.000\s?|[\d\.]+\s?VNĐ|[\d\.]+\s?VND|\d+\.\d{3}\s?đ|\d+\.\d{3}\s?₫|₫\s?\d+\.\d{3}|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',desc_content)
                            #if c == 3: col = re.sub('\d+\.\d+|d\d+|\d+d|đ[\d\s\.\,]|[\d\s\.\,]đ|[\d\s\.\,]+₫|₫[\d\.\,]+|\d+k|\d+K|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',col)
                            _ = worksheet.cell(column=c+1, row=r+rowid+i, value=col)

                rowid += len(sizes) - 1    
            elif len(colors) > 0:
                for ic,color in enumerate(colors):
                    size = ''
                    #colors = sel.xpath('./@title').extract_first()
                    #image = sel.xpath('./input/@onclick').re('changeImageShop\(([^\)]+)')
                    #image = re.sub('\'','',''.join(image).split(',')[-1])
                    #image = sel.xpath('//h4[contains(text(),"Màu")]/ancestor::div[2]/ul/li/label/@data-image').extract_first() or image
                    for c, col in enumerate(row):
                        if col == 11: col = 'https://shopee.vn/%s%s'%(re.sub('\W+','-',row[3]),re.sub('.*(?=-i\.)','',col))
                        if c == 13: col = cate
                        if re.search('size',color,re.I):
                            size = color
                            color = ''
                        if c == 9: col = color
                        if c == 1: col=max(eval(row[1] or '0'),eval(row[2] or '0'))
                        if c == 2: col=min(eval(row[1] or '0'),eval(row[2] or '0'))  
                        if c ==4: col = ''
                        if c == 7: col= brand
                        if c == 8: col = size
                        if c == 5: col = re.sub('|[\d\.]+\.000\s?|[\d\.]+\s?VNĐ|[\d\.]+\s?VND|\d+\.\d+\s?đ|\d+\.\d+\s?₫|₫\s?\d+\.\d+|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',col)    
                        if c== 6: col = re.sub('[\d\.]+\.000\s?|[\d\.]+\s?VNĐ|[\d\.]+\s?VND|\d+\.\d{3}\s?đ|\d+\.\d{3}\s?₫|₫\s?\d+\.\d{3}|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',desc_content)
                        #if c == 3: col = re.sub('\d+\.\d+|d\d+|\d+d|đ[\d\s\.\,]|[\d\s\.\,]đ|[\d\s\.\,]+₫|₫[\d\.\,]+|\d+k|\d+K|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',col)
                        if c == 10:
                            images1 = [re.sub('_tn$','',x) for x in eval(re.sub('background-image: url\(\"|\);[^\"]+\"','',re.sub(r'\\','',images_raw))) if x.strip()]
                            col = len(eval(row[10]))>0 and validator_image(images1,folder) or None
                            for ii,x in enumerate(len(images1) > 5 and images1[2:] or images1[1:]):
                                x=re.sub('_tn$','',x)
                                _ = worksheet.cell(column=c+1+5+ii, row=r+rowid+ic, value=x)
                        #if c == 10 and col is not None:  url_check(col) and True or print(col)
                        _ = worksheet.cell(column=c+1, row=r+rowid+ic, value=col)
                rowid += len(colors) - 1

            else:
                for c, col in enumerate(row):
                    if col == 11: col = 'https://shopee.vn/%s%s'%(re.sub('\W+','-',row[3]),re.sub('.*(?=-i\.)','',col))
                    if c == 13: col = cate
                    if c == 1: col=max(eval(row[1] or '0'),eval(row[2] or '0'))
                    if c == 2: col=min(eval(row[1] or '0'),eval(row[2] or '0')) 
                    #if c == 10: 
                    #    col=len(eval(col))>0 and validator_image(eval(col),folder) or None
                    #    for ii,x in enumerate(eval(row[10])):
                    #        x=re.sub('_tn$','',x)
                    #        _ = worksheet.cell(column=c+1+5+ii, row=r+rowid, value=x)
                    if c == 10:
                        images1 = [re.sub('_tn$','',x) for x in eval(re.sub('background-image: url\(\"|\);[^\"]+\"','',re.sub(r'\\','',images_raw))) if x.strip()]
                        col = len(eval(row[10]))>0 and validator_image(images1,folder) or None
                        for ii,x in enumerate(len(images1) > 5 and images1[2:] or images1[1:]):
                            x=re.sub('_tn$','',x)
                            _ = worksheet.cell(column=c+1+5+ii, row=r+rowid, value=x)
                    #if c == 10 and col is not None:  url_check(col) and True or print(col)
                    #if c == 11: col = html.unescape(row[11])
                    if c == 9: col = row[4]
                    if c == 8: col = row[13]
                    if c ==7: col = brand
                    if c== 6: col = re.sub('[\d\.]+\.000\s?|[\d\.]+\s?VNĐ|[\d\.]+\s?VND|\d+\.\d{3}\s?đ|\d+\.\d{3}\s?₫|₫\s?\d+\.\d{3}|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',desc_content)
                    if c ==4: col = ''
                    if c == 5: col = re.sub('|[\d\.]+\.000\s?|[\d\.]+\s?VNĐ|[\d\.]+\s?VND|\d+\.\d+\s?đ|\d+\.\d+\s?₫|₫\s?\d+\.\d+|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',col)
                    #if c == 3: col = re.sub('\(\)','',re.sub('\d+\.\d+\s?₫|₫\s?\d+\.\d+|\d+k\/?|\d+K\/?|GIÁ([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:\s?[\d\.,]+|Giá([\w\s]+)?:?\s?[\d\.,]+','',col))    
                    _ = worksheet.cell(column=c+1, row=r+rowid, value=col)
            

    # Auto-fit the column widths
    for column in worksheet.columns:
        max_length = 0
        column_name = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 70:
            adjusted_width = 70
        worksheet.column_dimensions[column_name].width = adjusted_width

    # Save the workbook as a BytesIO object
    output = io.BytesIO()
    workbook.save(output)

    # Set the content type and filename for the response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f"attachment; filename=[{filter1}].xlsx"

    return response

@socketio.on('process_output')
def handle_process_output(output):
    emit('update_textarea', output)

@app.route('/run-domain')
def run_domain():
    import paramiko

    domain_netloc = request.args.get('domain_netloc', '').strip()
    tool_name = request.args.get('tool_name', '').strip()

    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect("192.168.101.27", username="ml", password="1234")
    stdin, stdout, stderr = ssh.exec_command('/home/ml/lazada/skus_script.sh %s %s'%(tool_name,domain_netloc))

    # Wait for the command to complete
    exit_status = stdout.channel.recv_exit_status()

    # Return a message to indicate that the function has finished
    if exit_status == 0:
        return jsonify({'message': 'Done!'})
    else:
        return jsonify({'message': 'Error!'})

if __name__ == '__main__':
    app.run(debug=True)